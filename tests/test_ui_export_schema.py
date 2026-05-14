import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import fpt_einvoice_exporter as mod


class UiExportSchemaTests(unittest.TestCase):
    def _sample_row(self):
        return {
            "type": "01/MTT",
            "type_label": "Hóa đơn GTGT từ máy tính tiền",
            "idt": "2026-05-12T16:00:35.000Z",
            "inc": 9971061,
            "form": "1",
            "serial": "C26MHN",
            "seq": "00158238",
            "status": 7,
            "status_received": 8,
            "ma_cqthu": "M1-26-DEMO-00000000001",
            "stax": "0100123456",
            "sname": "CÔNG TY DEMO",
            "btax": None,
            "bcode": "KH001",
            "bname": "Khách demo",
            "buyer": "Khách demo",
            "baddr": "Địa chỉ demo",
            "bmail": "demo@example.com",
            "btel": "0900000000",
            "idnumber": "012345678901",
            "sum": "659258.0000",
            "vat": "52742.0000",
            "total": "712000.0000",
            "curr": "VND",
            "exrt": "1.00",
            "sec": "58z26llm25",
            "ou": 1,
            "uc": "API_HN",
            "ic": "gtgt_24639042_1",
            "adjdes": None,
            "cde": None,
            "class": "0",
            "business_class": None,
            "dtl_invs_status": None,
            "sec_dtl": None,
            "canrdt": None,
            "canref": None,
            "canrea": None,
            "adjrdt": None,
            "adjref": None,
            "adjrea": None,
            "minutes_status_sign": None,
            "code_minutes_sign": None,
            "note": None,
            "maildt": "2",
            "converted_inv": "",
            "id_batch": 44834,
            "dt": "2026-05-12T16:05:00.000Z",
            "org": {"on": "CÔNG TY DEMO"},
        }

    def test_export_workbook_uses_webapp_header_order_for_invoice_sheet(self):
        expected_headers = [
            "ID",
            "Ngày HĐ",
            "Mẫu số",
            "Ký hiệu",
            "Số HĐ",
            "Trạng thái",
            "Trạng thái CQT",
            "Mã CQT cấp",
            "MST",
            "Tên KH",
            "Người mua",
            "Địa chỉ",
            "SĐT người mua",
            "CCCD người mua",
            "Tổng tiền",
            "VAT",
            "Tổng cộng",
            "Loại tiền",
            "Tỷ giá",
            "Mã TC",
            "Email",
            "Đơn vị bán",
            "Người lập",
            "IC",
            "Thay thế/Điều chỉnh",
            "Bị thay thế/Bị điều chỉnh",
            "Phân loại nghiệp vụ",
            "ĐC/TT nhiều HĐ",
            "Mã bảng kê ĐC/TT",
            "Ngày hủy",
            "Số văn bản của HD hủy",
            "Lý do hủy",
            "Ngày VB",
            "Số văn bản của HD điều chỉnh",
            "Lý do",
            "TT ký biên bản",
            "Mã BB",
            "Mã KH",
            "Ghi chú",
            "Gửi mail",
            "Chuyển đổi",
            "ID gói gửi CQT",
            "Ngày gửi dữ liệu",
        ]

        row = self._sample_row()
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "out.xlsx"
            mod.export_workbook(
                output,
                [row],
                {"01/MTT": [row]},
                {"total_rows": 1, "counts": {"01/MTT": 1}},
            )
            wb = load_workbook(output, read_only=True)
            ws = wb["invoices_all"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        self.assertEqual(headers, expected_headers)
        self.assertNotIn("type_label", headers)
        self.assertNotIn("sname", headers)
        self.assertNotIn("stax", headers)
        self.assertNotIn("org", headers)

    def test_export_workbook_formats_invoice_values_like_webapp(self):
        row = self._sample_row()
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "out.xlsx"
            mod.export_workbook(
                output,
                [row],
                {"01/MTT": [row]},
                {"total_rows": 1, "counts": {"01/MTT": 1}},
            )
            wb = load_workbook(output, read_only=True)
            ws = wb["invoices_all"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            values = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]

        data = dict(zip(headers, values))
        self.assertEqual(data["Ngày HĐ"], "12/05/2026")
        self.assertEqual(data["Trạng thái"], "Đã phát hành")
        self.assertEqual(data["Trạng thái CQT"], "Kiểm tra hợp lệ")
        self.assertEqual(data["Đơn vị bán"], "CÔNG TY DEMO")
        self.assertEqual(data["Tổng tiền"], 659258)
        self.assertEqual(data["VAT"], 52742)
        self.assertEqual(data["Tổng cộng"], 712000)
        self.assertEqual(data["Tỷ giá"], 1)
        self.assertEqual(data["TT ký biên bản"], "Chưa ký")
        self.assertEqual(data["Gửi mail"], "Chưa gửi")
        self.assertEqual(data["Ngày gửi dữ liệu"], "12/05/2026")

    def test_export_workbook_sorts_invoices_all_by_datetime_then_id_desc(self):
        newer = self._sample_row()
        older = dict(self._sample_row())
        older["inc"] = 9969000
        older["idt"] = "2026-05-11T16:00:35.000Z"
        older["dt"] = "2026-05-11T16:05:00.000Z"
        older["seq"] = "00158000"

        same_time_lower_id = dict(self._sample_row())
        same_time_lower_id["inc"] = 9971000
        same_time_lower_id["seq"] = "00158001"

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "out.xlsx"
            mod.export_workbook(
                output,
                [older, same_time_lower_id, newer],
                {"01/MTT": [older, same_time_lower_id, newer]},
                {"total_rows": 3, "counts": {"01/MTT": 3}},
            )
            wb = load_workbook(output, read_only=True)
            ws = wb["invoices_all"]
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            id_idx = headers.index("ID")
            ids = [row[id_idx] for row in ws.iter_rows(min_row=2, max_row=4, values_only=True)]

        self.assertEqual(ids, [9971061, 9971000, 9969000])


if __name__ == "__main__":
    unittest.main()
