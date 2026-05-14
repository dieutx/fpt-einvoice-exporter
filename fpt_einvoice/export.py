import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from .constants import (
    CONVERTED_INV_LABELS,
    INVOICE_STATUS_LABELS,
    KNOWN_TYPE_LABELS,
    MAIL_STATUS_LABELS,
    MINUTES_SIGN_STATUS_LABELS,
    STATUS_RECEIVED_LABELS,
    UI_EXPORT_COLUMNS,
)
from .formatting import display_date, display_number, localize_iso, lookup_label


def resolve_ou_display(row: dict[str, Any]) -> Any:
    if row.get("ou_display") not in (None, ""):
        return row.get("ou_display")
    if row.get("sname") not in (None, ""):
        return row.get("sname")
    org = row.get("org")
    if isinstance(org, dict) and org.get("on"):
        return org.get("on")
    return row.get("ou", "")


def resolve_business_class(row: dict[str, Any]) -> Any:
    if row.get("business_class") not in (None, ""):
        return row.get("business_class")
    raw_class = row.get("class")
    if raw_class in (None, "", 0, "0"):
        return ""
    return raw_class


def resolve_dtl_invs_status(row: dict[str, Any]) -> Any:
    if row.get("dtl_invs_status") not in (None, ""):
        return row.get("dtl_invs_status")
    if not row.get("sec_dtl"):
        return ""
    if row.get("is_lock") in (1, "1", True):
        return "Chưa xử lý"
    return "Đã xử lý"


def resolve_cancel_date(row: dict[str, Any]) -> str:
    cancel_date = row.get("canrdt") or row.get("endtime")
    cde = str(row.get("cde") or "")
    if str(row.get("status")) == "4" and row.get("cid") and "Bị thay thế" in cde:
        cancel_date = ""
    return display_date(cancel_date)


def build_ui_export_row(row: dict[str, Any]) -> dict[str, Any]:
    values = {
        "inc": row.get("inc", ""),
        "idt": display_date(row.get("idt")),
        "form": row.get("form", ""),
        "serial": row.get("serial", ""),
        "seq": row.get("seq", ""),
        "status": lookup_label(row.get("status"), INVOICE_STATUS_LABELS),
        "status_received": lookup_label(row.get("status_received"), STATUS_RECEIVED_LABELS),
        "ma_cqthu": row.get("ma_cqthu", ""),
        "btax": row.get("btax", "") or "",
        "bname": row.get("bname", "") or "",
        "buyer": row.get("buyer", "") or "",
        "baddr": row.get("baddr", "") or "",
        "btel": row.get("btel", "") or "",
        "idnumber": row.get("idnumber") or row.get("passport_number") or "",
        "sum": display_number(row.get("sum")),
        "vat": display_number(row.get("vat")),
        "total": display_number(row.get("total")),
        "curr": row.get("curr", "") or "",
        "exrt": display_number(row.get("exrt")),
        "sec": row.get("sec", "") or "",
        "bmail": row.get("bmail", "") or "",
        "ou": resolve_ou_display(row),
        "uc": row.get("uc", "") or "",
        "ic": row.get("ic", "") or "",
        "adjdes": row.get("adjdes", "") or "",
        "cde": row.get("cde", "") or "",
        "business_class": resolve_business_class(row),
        "dtl_invs_status": resolve_dtl_invs_status(row),
        "sec_dtl": row.get("sec_dtl", "") or "",
        "canrdt": resolve_cancel_date(row),
        "canref": row.get("canref", "") or "",
        "canrea": row.get("canrea", "") or "",
        "adjrdt": display_date(row.get("adjrdt")),
        "adjref": row.get("adjref", "") or "",
        "adjrea": row.get("adjrea", "") or "",
        "minutes_status_sign": lookup_label(
            "null" if row.get("minutes_status_sign") is None else row.get("minutes_status_sign"),
            MINUTES_SIGN_STATUS_LABELS,
        ),
        "code_minutes_sign": row.get("code_minutes_sign", "") or "",
        "bcode": row.get("bcode", "") or "",
        "note": row.get("note", "") or "",
        "maildt": lookup_label(row.get("maildt"), MAIL_STATUS_LABELS),
        "converted_inv": lookup_label(row.get("converted_inv"), CONVERTED_INV_LABELS),
        "id_batch": row.get("id_batch", "") or "",
        "dt": display_date(row.get("dt")),
    }
    return {header: values.get(field_id, "") for field_id, header in UI_EXPORT_COLUMNS}


def normalize_value(value: Any) -> Any:
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        return json.dumps(value, ensure_ascii=False)
    return value


def flatten_invoice(row: dict[str, Any], type_label: str) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        out[key] = normalize_value(value)
    out["type_label"] = type_label
    out["idt_local"] = localize_iso(row.get("idt"))
    return out


def sheet_name_for_type(code: str, label: str) -> str:
    base = f"{code} {label}".replace("/", "-")
    bad = set('[]:*?/\\')
    cleaned = "".join("-" if c in bad else c for c in base)
    return cleaned[:31]


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > length:
                length = len(val)
        ws.column_dimensions[col].width = min(max(length + 2, 10), 40)


def write_sheet(ws, rows: Iterable[dict[str, Any]], columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _sort_invoice_rows(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    def sort_key(row: dict[str, Any]) -> tuple[str, int]:
        raw_inc = row.get("inc")
        try:
            inc_value = int(raw_inc)
        except (TypeError, ValueError):
            inc_value = -1
        return (str(row.get("idt") or ""), inc_value)

    return sorted(rows, key=sort_key, reverse=True)


def export_workbook(
    output_xlsx: Path,
    all_rows: list[dict[str, Any]],
    by_type: dict[str, list[dict[str, Any]]],
    metadata: dict[str, Any],
) -> None:
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "metadata"
    ws_meta.append(["key", "value"])
    for cell in ws_meta[1]:
        cell.font = Font(bold=True)
    for key, value in metadata.items():
        ws_meta.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    autosize_columns(ws_meta)
    ws_meta.freeze_panes = "A2"

    ws_summary = wb.create_sheet("summary")
    ws_summary.append(["type_code", "type_label", "count"])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
    for code, rows in by_type.items():
        ws_summary.append([code, KNOWN_TYPE_LABELS.get(code, code), len(rows)])
    ws_summary.append(["TOTAL", "Tất cả", len(all_rows)])
    autosize_columns(ws_summary)
    ws_summary.freeze_panes = "A2"

    ui_headers = [header for _, header in UI_EXPORT_COLUMNS]
    sorted_all_rows = _sort_invoice_rows(all_rows)
    ui_all_rows = [build_ui_export_row(row) for row in sorted_all_rows]
    ws_all = wb.create_sheet("invoices_all")
    write_sheet(ws_all, ui_all_rows, ui_headers)

    for code, rows in by_type.items():
        ws = wb.create_sheet(sheet_name_for_type(code, KNOWN_TYPE_LABELS.get(code, code)))
        sorted_type_rows = _sort_invoice_rows(rows)
        ui_rows = [build_ui_export_row(row) for row in sorted_type_rows]
        write_sheet(ws, ui_rows, ui_headers)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
